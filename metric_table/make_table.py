from __future__ import annotations

import json
import math
import re
from collections import defaultdict
from pathlib import Path
from statistics import fmean
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT = SCRIPT_DIR / "metrics_tables.xlsx"


# ---------------------------------------------------------------------
# JSON / numeric helpers
# ---------------------------------------------------------------------

def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return 1.0 if value else 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    try:
        s = str(value).strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def mean_or_none(values: List[Any]) -> Optional[float]:
    nums = []
    for v in values:
        x = to_float(v)
        if x is not None:
            nums.append(x)
    if not nums:
        return None
    return float(fmean(nums))


def display_percent(value: Optional[float]) -> Optional[float]:
    """
    Convert a 0..1 rate to 0..100 for display.
    If the value already looks like a percentage, leave it unchanged.
    """
    if value is None:
        return None
    if value <= 1.5:
        return value * 100.0
    return value


def pick_first(record: Dict[str, Any], keys: Tuple[str, ...]) -> Any:
    for k in keys:
        if k in record and record[k] not in (None, "", [], {}):
            return record[k]
    return None


# ---------------------------------------------------------------------
# File discovery
# ---------------------------------------------------------------------

def find_metrics_files(root: Path) -> List[Path]:
    files = [
        root / "Reasoning vs Non-reasoning" / "metrics_new.json",
        root / "Runs" / "Untitled" / "metrics_new.json",
        root / "Small vs Large" / "metrics_new.json",
    ]
    return [p for p in files if p.exists()]


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]+", "_", name)
    name = re.sub(r"\s+", "_", name).strip("_")
    return name[:31] if len(name) > 31 else name


def unique_sheet_name(existing: set[str], base: str) -> str:
    name = safe_sheet_name(base)
    if name not in existing:
        existing.add(name)
        return name

    i = 2
    while True:
        suffix = f"_{i}"
        trimmed = name[: 31 - len(suffix)]
        candidate = f"{trimmed}{suffix}"
        if candidate not in existing:
            existing.add(candidate)
            return candidate
        i += 1


# ---------------------------------------------------------------------
# Aggregation
# ---------------------------------------------------------------------

def group_by_model_pair(records: List[Dict[str, Any]]) -> Dict[Tuple[str, str], List[Dict[str, Any]]]:
    grouped: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
    for rec in records:
        model1 = str(rec.get("Model1", "")).strip()
        model2 = str(rec.get("Model2", "")).strip()
        grouped[(model1, model2)].append(rec)
    return grouped


def aggregate_rows(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped = group_by_model_pair(records)
    output_rows: List[Dict[str, Any]] = []

    for (model1, model2), rows in grouped.items():
        nc_vals = [pick_first(r, ("NC_Rate", "NC")) for r in rows]
        ar_vals = [pick_first(r, ("AR_Rate", "AR")) for r in rows]
        fr_vals = [pick_first(r, ("FR_Rate", "FR")) for r in rows]

        fav_si_vals = [pick_first(r, ("ArgsPerSample_Fav_SI", "Avg_ArgsPerSample_Fav_SI")) for r in rows]
        opp_si_vals = [pick_first(r, ("ArgsPerSample_Opp_SI", "Avg_ArgsPerSample_Opp_SI")) for r in rows]
        fav_dbt_vals = [pick_first(r, ("ArgsPerSample_Fav_Dbt", "Avg_ArgsPerSample_Fav_Dbt")) for r in rows]
        opp_dbt_vals = [pick_first(r, ("ArgsPerSample_Opp_Dbt", "Avg_ArgsPerSample_Opp_Dbt")) for r in rows]

        agr_vals = [pick_first(r, ("Agr_Rate", "Agr")) for r in rows]
        disagr_vals = [pick_first(r, ("DisAgr_Rate", "DisAgr")) for r in rows]
        ign_vals = [pick_first(r, ("Ign_Rate", "Ign")) for r in rows]

        verb_fav_vals = [pick_first(r, ("Verbosity_Fav",)) for r in rows]
        verb_opp_vals = [pick_first(r, ("Verbosity_Opp",)) for r in rows]

        steps_vals = [pick_first(r, ("Avg_Steps", "Steps")) for r in rows]

        row = {
            "Model Pair": f"{model1} vs {model2}",
            "NC (%)": display_percent(mean_or_none(nc_vals)),
            "AR (%)": display_percent(mean_or_none(ar_vals)),
            "FR (%)": display_percent(mean_or_none(fr_vals)),
            "Fav SI": mean_or_none(fav_si_vals),
            "Opp SI": mean_or_none(opp_si_vals),
            "Fav Dbt": mean_or_none(fav_dbt_vals),
            "Opp Dbt": mean_or_none(opp_dbt_vals),
            "Agr (%)": display_percent(mean_or_none(agr_vals)),
            "DisAgr (%)": display_percent(mean_or_none(disagr_vals)),
            "Ign (%)": display_percent(mean_or_none(ign_vals)),
            "Fav": mean_or_none(verb_fav_vals),
            "Opp": mean_or_none(verb_opp_vals),
            "# Steps": mean_or_none(steps_vals),
            "_count": len(rows),
        }
        output_rows.append(row)

    output_rows.sort(key=lambda r: str(r["Model Pair"]).lower())
    return output_rows


# ---------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------

def set_border(cell, border: Border) -> None:
    cell.border = border


def style_sheet(ws, rows: List[Dict[str, Any]], source_label: str) -> None:
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name="Times New Roman", size=11, bold=True)
    body_font = Font(name="Times New Roman", size=11)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:N1")
    title = ws["A1"]
    title.value = source_label
    title.font = Font(name="Times New Roman", size=12, bold=True)
    title.alignment = left

    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    ws.merge_cells("A2:A3")
    ws["A2"] = "Model Pair"
    ws["A2"].font = header_font
    ws["A2"].alignment = center
    set_border(ws["A2"], border)

    groups = [
        ("SI vs Dbt", "B2:D2"),
        ("<args/sample>", "E2:H2"),
        ("Argument Response in Dbt", "I2:K2"),
        ("Verbosity", "L2:M2"),
        ("# Steps", "N2:N3"),
    ]
    for label, rng in groups:
        ws.merge_cells(rng)
        c = ws[rng.split(":")[0]]
        c.value = label
        c.font = header_font
        c.alignment = center
        set_border(c, border)

    subheaders = {
        "B3": "NC (%)", "C3": "AR (%)", "D3": "FR (%)",
        "E3": "Fav SI", "F3": "Opp SI", "G3": "Fav Dbt", "H3": "Opp Dbt",
        "I3": "Agr (%)", "J3": "DisAgr (%)", "K3": "Ign (%)",
        "L3": "Fav", "M3": "Opp",
    }
    for cell_ref, label in subheaders.items():
        cell = ws[cell_ref]
        cell.value = label
        cell.font = header_font
        cell.alignment = center
        set_border(cell, border)

    start_row = 4
    for r_idx, row in enumerate(rows, start=start_row):
        values = [
            row["Model Pair"], row["NC (%)"], row["AR (%)"], row["FR (%)"],
            row["Fav SI"], row["Opp SI"], row["Fav Dbt"], row["Opp Dbt"],
            row["Agr (%)"], row["DisAgr (%)"], row["Ign (%)"],
            row["Fav"], row["Opp"], row["# Steps"],
        ]

        for c_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = body_font
            cell.alignment = left if c_idx == 1 else center
            set_border(cell, border)

            if c_idx != 1 and value is not None:
                cell.number_format = "0.0"

    widths = {1:34,2:10,3:10,4:10,5:10,6:10,7:10,8:10,9:12,10:12,11:10,12:10,13:10,14:10}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def build_workbook(root: Path, output_path: Path) -> None:
    metrics_files = find_metrics_files(root)
    if not metrics_files:
        raise FileNotFoundError(f"No metrics_new.json files found under: {root}")

    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()

    for metrics_file in metrics_files:
        records = load_json(metrics_file)
        if not isinstance(records, list):
            continue

        rows = aggregate_rows(records)

        rel = metrics_file.relative_to(root)

        # ===== ONLY CHANGE =====
        parent_name = metrics_file.parent.name
        if parent_name == "Reasoning vs Non-reasoning":
            sheet_name = unique_sheet_name(used_names, "Reasoning_vs_NonReasoning")
        elif parent_name == "Untitled":
            sheet_name = unique_sheet_name(used_names, "Runs_Untitled")
        elif parent_name == "Small vs Large":
            sheet_name = unique_sheet_name(used_names, "Small_vs_Large")
        else:
            sheet_name = unique_sheet_name(used_names, parent_name)
        # =======================

        print(f"Creating sheet for: {metrics_file}")

        ws = wb.create_sheet(title=sheet_name)
        source_label = f"Source: {rel.as_posix()}"
        style_sheet(ws, rows, source_label)

    wb.save(output_path)


def main() -> None:
    root = Path.cwd()
    output_path = DEFAULT_OUTPUT
    build_workbook(root, output_path)
    print(f"Wrote workbook to: {output_path}")


if __name__ == "__main__":
    main()